#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
FES IDA-1 Compiler Module
Calculates IDA-1 adjustment bids by comparing IDA-1 forecast to D-1 forecast
Adjustment = IDA-1 Total - D-1 Total (horizontally summed)
"""

import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl.styles import Font, PatternFill, Alignment
from zeep import Client
from zeep.settings import Settings


class IDA1BidCompiler:
    def __init__(self):
        """Initialize the IDA-1 compiler"""
        self.cwd = Path.cwd()
        self.d1_forecast_df = None
        self.ida1_forecast_df = None
        self.adjustment_df = None
    
    def download_ida1_forecast(self, trading_date_str, upload_sql=False):
        """
        Download IDA-1 forecast from Meteologica API and save
        (Same as D-1 but saves with IDA-1 suffix)
        
        Args:
            trading_date_str: Trading date in DD/MM/YYYY format
            upload_sql: Whether to upload to SQL (default: False)
        
        Returns:
            DataFrame with forecast data
        """
        print(f"\n[IDA1] Downloading IDA-1 forecast from Meteologica API...")
        
        date_obj_1 = datetime.strptime(trading_date_str, "%d/%m/%Y") - timedelta(days=1)
        date_obj_2 = datetime.strptime(trading_date_str, "%d/%m/%Y")

        y1, m1, d1 = date_obj_1.year, date_obj_1.month, date_obj_1.day
        y2, m2, d2 = date_obj_2.year, date_obj_2.month, date_obj_2.day

        settings = Settings(strict=False)
        wsdl = "https://webservice.meteologica.com/api/wsdl/MeteologicaDataExchangeService.wsdl"
        client = Client(wsdl=wsdl)

        login_req_type = client.get_type('ns0:LoginReq')
        login_req = login_req_type(
            username="Flogas",
            password="uc:DF824"
        )
        login_response = client.service.login(request=login_req)
        session_token = login_response.header.sessionToken

        forecast_req_type = client.get_type('ns0:GetForecastMultiReq')
        req = forecast_req_type(
            header={'sessionToken': session_token},
            variableId='prod',
            predictorId='aggregated',
            fromDate=datetime(y1, m1, d1, 23, 0, 0).isoformat(),
            toDate=datetime(y2, m2, d2, 22, 30, 0).isoformat(),
            granularity='30',
            percentiles='50',
            facilitiesId=[])

        forecast_response = client.service.getForecastMulti(request=req)

        all_timestamps = set()
        facility_data = {}

        for facility in forecast_response.facilitiesForecastData.item:
            facility_id = facility.facilityId
            facility_data[facility_id] = {}

            forecast_pairs = facility.forecastData.split(':')
            start_idx = 1 if forecast_pairs[0] == '' else 0

            for i in range(start_idx, len(forecast_pairs)):
                parts = forecast_pairs[i].split('~')
                if len(parts) >= 2:
                    timestamp = int(parts[0])
                    value = float(parts[1])
                    dt_obj = datetime.utcfromtimestamp(timestamp)
                    all_timestamps.add(dt_obj)
                    facility_data[facility_id][dt_obj] = value / 1000  # Convert to MW

        df = pd.DataFrame(
            index=pd.to_datetime(sorted(all_timestamps)),
            columns=[
                'Meteo ROI (MW)', 'Meteo NI (MW)', 'Meteo TB (MW)',
                'Meteo CK (MW)', 'Meteo LD (MW)', 'Meteo CD (MW)', 'Meteo DT (MW)', 
                'Meteo MUR (MW)', 'Meteo S1 (MW)', 'Meteo S2 (MW)'
            ]
        )

        column_mapping = {
            'Vayu_Cluster1': 'Meteo ROI (MW)',
            'Vayu_Cluster2': 'Meteo NI (MW)',
            'Vayu_402050': 'Meteo TB (MW)',
            'Vayu_GU_402280': 'Meteo CK (MW)',
            'Flogas-solar_0587': 'Meteo LD (MW)',
            'Vayu_0275': 'Meteo CD (MW)',
            'Flogas-solar_0378__': 'Meteo DT (MW)',
            'Vayu_GEN_504260': 'Meteo MUR (MW)',
            'Flogas-solar_0670': 'Meteo S1 (MW)',
            'Flogas-solar_0684': 'Meteo S2 (MW)'
        }

        for facility_id, col_name in column_mapping.items():
            if facility_id in facility_data:
                df[col_name] = [facility_data[facility_id].get(ts, 0) for ts in df.index]

        df = df.reset_index().rename(columns={'index': 'time'})
        df['time'] = df['time'] + pd.to_timedelta('1 hour')
        df['time'] = df['time'].dt.strftime('%d/%m/%Y %H:%M')

        # Process forecast data (add nonwind, self-forecast, etc.)
        # Import process_forecast_data from master script
        from FES_MasterScript_PRODUCTION import process_forecast_data
        df = process_forecast_data(trading_date_str, df)
        df.rename(columns={"time": "DateTime"}, inplace=True)

        # Set proper timestamps
        input_date_obj = datetime.strptime(trading_date_str, "%d/%m/%Y")
        previous_day = input_date_obj - timedelta(days=1)
        timestamps = []
        for minute in [0, 30]:
            timestamp = previous_day.replace(hour=23, minute=minute)
            timestamps.append(timestamp)
        for hour in range(0, 23):
            for minute in [0, 30]:
                timestamp = input_date_obj.replace(hour=hour, minute=minute)
                timestamps.append(timestamp)

        df['DateTime'] = timestamps
        df['DateTime'] = df['DateTime'].dt.strftime('%d/%m/%Y %H:%M')

        df.loc[:, df.columns != 'DateTime'] = df.loc[:, df.columns != 'DateTime'].round(1)

        # === SAVE AS IDA-1 ===
        date_obj = datetime.strptime(trading_date_str, "%d/%m/%Y")
        year = str(date_obj.year)
        month = date_obj.strftime("%B")
        
        output_path = Path(rf"I:\Daily Generation Forecasts\Daily Generation to Submit\{year}\{month}")
        output_path.mkdir(parents=True, exist_ok=True)
        
        file_name = f"Generation Forecast {date_obj.strftime('%d.%m.%Y')} IDA-1.xlsx"
        full_path = output_path / file_name
        df.to_excel(full_path, index=False)
        print(f"[IDA1] OK Forecast saved: {file_name}")
        
        # === UPLOAD TO FABRIC SQL (ONLY IF ENABLED) ===
        # IDA-1 is not D-1, so it goes to Generation_D_Minus_X table
        if upload_sql:
            try:
                from FES_MasterScript_PRODUCTION import upload_to_fabric
                print(f"[IDA1] Uploading to Fabric SQL (Generation_D_Minus_X table)...")
                upload_success = upload_to_fabric(df, file_name)  # File name contains "IDA-1", will route to D_Minus_X
                if upload_success:
                    print(f"[IDA1] OK SQL upload complete")
                else:
                    print(f"[IDA1] WARNING: SQL upload failed")
            except Exception as e:
                print(f"[IDA1] WARNING: SQL upload failed: {str(e)}")
                print(f"[IDA1] Continuing without SQL upload...")
        else:
            print(f"[IDA1] SKIP SQL upload (disabled)")

        logout_req_type = client.get_type('ns0:LogoutReq')(
            header={'sessionToken': session_token})
        client.service.logout(request=logout_req_type)

        return df
        
    def load_d1_forecast(self, trading_date_str):
        """Load D-1 forecast from morning"""
        date_obj = datetime.strptime(trading_date_str, "%d/%m/%Y")
        year = str(date_obj.year)
        month = date_obj.strftime("%B")
        day_str = date_obj.strftime("%d.%m.%Y")
        
        forecast_dir = Path(rf"I:\Daily Generation Forecasts\Daily Generation to Submit\{year}\{month}")
        d1_file = forecast_dir / f"Generation Forecast {day_str} D-1.xlsx"
        
        print(f"[IDA1] Loading D-1 forecast: {d1_file}")
        
        if not d1_file.exists():
            raise FileNotFoundError(f"D-1 forecast not found: {d1_file}\nPlease run D-1 compilation first (morning)")
        
        self.d1_forecast_df = pd.read_excel(d1_file)
        self.d1_forecast_df["DateTime"] = pd.to_datetime(self.d1_forecast_df["DateTime"], format="%d/%m/%Y %H:%M")
        print(f"[IDA1] OK D-1 forecast loaded: {len(self.d1_forecast_df)} periods")
        
    def load_forecasts(self, trading_date_str):
        """
        DEPRECATED - Use download_ida1_forecast and load_d1_forecast instead
        Load D-1 and IDA-1 forecast files
        
        Args:
            trading_date_str: Trading date in DD/MM/YYYY format
        """
        date_obj = datetime.strptime(trading_date_str, "%d/%m/%Y")
        year = str(date_obj.year)
        month = date_obj.strftime("%B")
        day_str = date_obj.strftime("%d.%m.%Y")
        
        forecast_dir = Path(rf"I:\Daily Generation Forecasts\Daily Generation to Submit\{year}\{month}")
        
        # Load D-1 forecast
        d1_file = forecast_dir / f"Generation Forecast {day_str} D-1.xlsx"
        print(f"\n[IDA1] Loading D-1 forecast: {d1_file}")
        
        if not d1_file.exists():
            raise FileNotFoundError(f"D-1 forecast not found: {d1_file}")
        
        self.d1_forecast_df = pd.read_excel(d1_file)
        self.d1_forecast_df["DateTime"] = pd.to_datetime(self.d1_forecast_df["DateTime"], format="%d/%m/%Y %H:%M")
        print(f"[IDA1] OK D-1 forecast loaded: {len(self.d1_forecast_df)} periods")
        
        # Load IDA-1 forecast
        ida1_file = forecast_dir / f"Generation Forecast {day_str} IDA-1.xlsx"
        print(f"[IDA1] Loading IDA-1 forecast: {ida1_file}")
        
        if not ida1_file.exists():
            raise FileNotFoundError(f"IDA-1 forecast not found: {ida1_file}\n" + 
                                   "Please generate IDA-1 forecast first (evening update)")
        
        self.ida1_forecast_df = pd.read_excel(ida1_file)
        self.ida1_forecast_df["DateTime"] = pd.to_datetime(self.ida1_forecast_df["DateTime"], format="%d/%m/%Y %H:%M")
        print(f"[IDA1] OK IDA-1 forecast loaded: {len(self.ida1_forecast_df)} periods")
        
    def calculate_adjustment(self):
        """Calculate IDA-1 adjustment = D-1 Total - IDA-1 Total
        
        Logic: When wind increases in IDA-1 forecast, we need to SELL MORE
        - D-1 was 100 MW, IDA-1 is 110 MW → Adjustment = 100 - 110 = -10 MW (sell 10 more)
        - D-1 was 100 MW, IDA-1 is 90 MW → Adjustment = 100 - 90 = +10 MW (buy back 10)
        """
        
        # Generation columns to sum
        gen_cols = ['Meteo ROI (MW)', 'Meteo NI (MW)', 'Meteo TB (MW)', 
                    'Meteo CK (MW)', 'Meteo LD (MW)', 'Meteo CD (MW)',
                    'Naïve Nonwind (MW)', 'Self-forecast (MW)', 'Meteo DT (MW)', 
                    'Meteo MUR (MW)']
        
        # Add S1-S25 columns if they exist
        for i in range(1, 26):
            col_name = f'Meteo S{i} (MW)'
            if col_name in self.d1_forecast_df.columns:
                gen_cols.append(col_name)
        
        # Calculate totals (sum horizontally)
        d1_cols = [col for col in gen_cols if col in self.d1_forecast_df.columns]
        ida1_cols = [col for col in gen_cols if col in self.ida1_forecast_df.columns]
        
        d1_total = self.d1_forecast_df[d1_cols].sum(axis=1)
        ida1_total = self.ida1_forecast_df[ida1_cols].sum(axis=1)
        
        # Calculate adjustment - CORRECT SIGN: D-1 minus IDA-1
        adjustment = d1_total - ida1_total
        
        # Create adjustment dataframe
        self.adjustment_df = pd.DataFrame({
            'DateTime': self.d1_forecast_df['DateTime'],
            'D1_Total': d1_total.round(1),
            'IDA1_Total': ida1_total.round(1),
            'Adjustment': adjustment.round(1)
        })
        
        print(f"\n[IDA1] Adjustment calculated:")
        print(f"  D-1 Total: {d1_total.sum():.1f} MWh")
        print(f"  IDA-1 Total: {ida1_total.sum():.1f} MWh")
        print(f"  Total Adjustment (D-1 - IDA-1): {adjustment.sum():.1f} MWh")
        print(f"  (Negative = Sell more, Positive = Buy back)")
        
        return self.adjustment_df
    
    def generate_ida1_bids(self):
        """Generate IDA-1 bid format (period, -150, 3000 structure)"""
        
        bid_data = []
        
        for i, row in self.adjustment_df.iterrows():
            adj = row['Adjustment']
            
            # IDA-1 bid structure: adjustment goes to both price columns
            bid_data.append({
                'period': i + 1,
                '-150': adj,
                '3000.00': adj
            })
        
        bid_df = pd.DataFrame(bid_data)
        
        # Add total row
        total_adj = self.adjustment_df['Adjustment'].sum()
        total_row = pd.DataFrame([{
            'period': '',
            '-150': total_adj.round(1),
            '3000.00': total_adj.round(1)
        }])
        
        bid_df = pd.concat([bid_df, total_row], ignore_index=True)
        
        return bid_df
    
    def create_ida1_excel_with_charts(self, trading_date_str):
        """Create single IDA Excel file with D-1, IDA-1, adjustment, bids, and embedded charts"""
        
        date_obj = datetime.strptime(trading_date_str, "%d/%m/%Y")
        year = str(date_obj.year)
        day_str = date_obj.strftime("%d.%m.%Y")
        
        # Output directory
        output_dir = Path(rf"I:\Intra-Day Process\ID-Auctions\ETS Upload\{year}")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Filename: IDA ETS {DD.MM.YYYY}.xlsx
        output_file = output_dir / f"IDA ETS {day_str}.xlsx"
        
        print(f"\n[IDA1] Creating IDA Excel: {output_file}")
        
        # Create temp chart directory
        temp_chart_dir = Path.cwd() / "temp_charts"
        temp_chart_dir.mkdir(exist_ok=True)
        
        times = self.adjustment_df['DateTime']
        d1_total = self.adjustment_df['D1_Total']
        ida1_total = self.adjustment_df['IDA1_Total']
        adjustment = self.adjustment_df['Adjustment']
        
        # Chart 1: D-1 vs IDA-1 comparison
        chart1_file = temp_chart_dir / "chart1.png"
        fig, ax = plt.subplots(figsize=(12, 6))
        
        ax.plot(times, d1_total, linewidth=2.5, color='#1f77b4', 
                label='D-1 Total', marker='o', markersize=3)
        ax.plot(times, ida1_total, linewidth=2.5, color='#ff7f0e', 
                label='IDA-1 Total', marker='o', markersize=3)
        
        ax.fill_between(times, d1_total, ida1_total, alpha=0.2, color='gray')
        
        ax.set_ylabel('MW', fontsize=12, fontweight='bold')
        ax.set_xlabel('Time', fontsize=12, fontweight='bold')
        ax.set_title(f'D-1 vs IDA-1 Forecast Comparison - {day_str}', 
                     fontsize=14, fontweight='bold', pad=15)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.legend(loc='upper right', fontsize=11)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
        plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        plt.savefig(chart1_file, dpi=150, bbox_inches='tight')
        plt.close()
        
        # Chart 2: Adjustment bars
        chart2_file = temp_chart_dir / "chart2.png"
        fig, ax = plt.subplots(figsize=(12, 6))
        
        colors = ['#2ca02c' if x < 0 else '#d62728' for x in adjustment]
        ax.bar(times, adjustment, width=0.015, color=colors, alpha=0.7, 
               label='Adjustment (D-1 - IDA-1)')
        ax.axhline(y=0, color='black', linestyle='-', linewidth=1, alpha=0.5)
        
        ax.set_xlabel('Time', fontsize=12, fontweight='bold')
        ax.set_ylabel('MW', fontsize=12, fontweight='bold')
        ax.set_title('IDA-1 Adjustment (Negative = Sell More, Positive = Buy Back)', 
                     fontsize=14, fontweight='bold', pad=15)
        ax.grid(True, alpha=0.3, linestyle='--')
        ax.legend(loc='upper right', fontsize=11)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
        plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        plt.savefig(chart2_file, dpi=150, bbox_inches='tight')
        plt.close()
        
        # Create Excel with all sheets
        from openpyxl.drawing.image import Image as OpenpyxlImage
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # Sheet 1: D-1 Forecast
            self.d1_forecast_df.to_excel(writer, sheet_name='D-1 Forecast', index=False)
            
            # Sheet 2: IDA-1 Forecast
            self.ida1_forecast_df.to_excel(writer, sheet_name='IDA-1 Forecast', index=False)
            
            # Sheet 3: Adjustment Calculation
            self.adjustment_df.to_excel(writer, sheet_name='Adjustment', index=False)
            
            # Sheet 4: IDA-1 Bids
            bid_df = self.generate_ida1_bids()
            bid_df.to_excel(writer, sheet_name='IDA1 Bids', index=False)
            
            # Sheet 5: Charts
            workbook = writer.book
            chart_sheet = workbook.create_sheet('Charts')
            
            # Add chart 1
            img1 = OpenpyxlImage(str(chart1_file))
            img1.anchor = 'A1'
            chart_sheet.add_image(img1)
            
            # Add chart 2 below chart 1
            img2 = OpenpyxlImage(str(chart2_file))
            img2.anchor = 'A35'  # Position below first chart
            chart_sheet.add_image(img2)
            
            # Format headers
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for sheet_name in ['D-1 Forecast', 'IDA-1 Forecast', 'Adjustment', 'IDA1 Bids']:
                ws = workbook[sheet_name]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
        
        # Clean up temp charts
        import shutil
        shutil.rmtree(temp_chart_dir)
        
        print(f"[IDA1] OK Excel with all data and charts saved")
        return output_file
    
    
    def upload_ida1_bids_to_sql(self, trading_date_str):
        """
        Upload IDA-1 bids to SQL table dbo.ida1_bids
        
        Table structure:
        - DateTime
        - IDA1_Bid_MW (the adjustment)
        - Meteo ROI (MW), Meteo NI (MW), etc. (all generation components)
        - Upload_Timestamp
        """
        try:
            import pyodbc
            import urllib.parse
            from sqlalchemy import create_engine
            
            print(f"[IDA1] Uploading IDA-1 bids to SQL (dbo.ida1_bids)...")
            
            # Prepare upload dataframe
            upload_df = self.ida1_forecast_df.copy()
            
            # Add IDA1_Bid_MW column (the adjustment)
            upload_df['IDA1_Bid_MW'] = self.adjustment_df['Adjustment (D-1 - IDA-1)']
            
            # Rename columns to match SQL table
            column_map = {
                'DateTime': 'DateTime',
                'Meteo ROI (MW)': 'Meteo ROI (MW)',
                'Meteo NI (MW)': 'Meteo NI (MW)',
                'Meteo TB (MW)': 'Meteo TB (MW)',
                'Meteo CK (MW)': 'Meteo CK (MW)',
                'Meteo LD (MW)': 'Meteo LD (MW)',
                'Meteo CD (MW)': 'Meteo CD (MW)',
                'Naïve Nonwind (MW)': 'Naïve Nonwind (MW)',
                'Self-forecast (MW)': 'Self-forecast (MW)',
                'Meteo DT (MW)': 'Meteo DT (MW)',
                'Meteo MUR (MW)': 'Meteo MUR (MW)',
                'Meteo S1 (MW)': 'Meteo S1 (MW)',
                'Meteo S2 (MW)': 'Meteo S2 (MW)'
            }
            
            # Select only columns that exist in the dataframe and are in the mapping
            available_cols = ['DateTime', 'IDA1_Bid_MW'] + [col for col in column_map.keys() if col in upload_df.columns and col != 'DateTime']
            upload_df = upload_df[available_cols]
            
            # Add upload timestamp
            upload_df['Upload_Timestamp'] = datetime.now()
            
            # Create connection
            server = 'g3hsqkj33hsejptu6vliyt5gny-6novrz7kmrcuriuozi2uqi5sy4.datawarehouse.fabric.microsoft.com'
            database = 'trading_data'
            
            conn_str = (
                f"Driver={{ODBC Driver 18 for SQL Server}};"
                f"Server={server};"
                f"Database={database};"
                f"Authentication=ActiveDirectoryInteractive;"
                f"Encrypt=yes;TrustServerCertificate=no;Connection Timeout=60;"
            )
            params = urllib.parse.quote_plus(conn_str)
            engine = create_engine(f"mssql+pyodbc:///?odbc_connect={params}")
            
            # Upload to SQL
            upload_df.to_sql('ida1_bids', engine, if_exists='append', index=False, method='multi')
            
            print(f"[IDA1] OK Uploaded {len(upload_df)} rows to dbo.ida1_bids")
            print(f"[IDA1] Columns: DateTime, IDA1_Bid_MW, Generation Components, Upload_Timestamp")
            
        except Exception as e:
            print(f"[IDA1] WARNING: SQL upload failed: {str(e)}")
            print(f"[IDA1] Continuing without SQL upload...")

    def run_ida1_compilation(self, trading_date_str, upload_sql=False):
        """
        Main workflow for IDA-1 bid compilation
        
        Args:
            trading_date_str: Trading date in DD/MM/YYYY format
            upload_sql: Whether to upload to SQL (default: False)
        
        Returns:
            Path to IDA Excel file
        """
        print("\n" + "="*70)
        print("IDA-1 BID COMPILATION WORKFLOW")
        print("="*70)
        print(f"Trading Day: {trading_date_str}")
        print(f"Execution Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"SQL Upload: {'ENABLED' if upload_sql else 'DISABLED'}")
        print("="*70 + "\n")
        
        # Step 1: Download IDA-1 forecast
        print("[STEP 1/4] Downloading IDA-1 forecast from Meteologica...")
        self.ida1_forecast_df = self.download_ida1_forecast(trading_date_str, upload_sql=upload_sql)
        
        # Step 2: Load D-1 forecast
        print("\n[STEP 2/4] Loading D-1 forecast (from morning)...")
        self.load_d1_forecast(trading_date_str)
        
        # Step 3: Calculate adjustment
        print("\n[STEP 3/4] Calculating adjustment (D-1 - IDA-1)...")
        self.calculate_adjustment()
        
        # Step 4: Upload IDA-1 bids to SQL (ONLY IF ENABLED)
        if upload_sql:
            print("\n[STEP 4/5] Uploading IDA-1 bids to SQL...")
            self.upload_ida1_bids_to_sql(trading_date_str)
        else:
            print("\n[STEP 4/5] SKIP Uploading IDA-1 bids to SQL (disabled)...")
        
        # Step 5: Create comprehensive Excel with all data and charts
        print("\n[STEP 5/5] Creating IDA Excel with all sheets and charts...")
        ida_excel = self.create_ida1_excel_with_charts(trading_date_str)
        
        print("\n" + "="*70)
        print("IDA-1 COMPILATION COMPLETE")
        print("="*70)
        print("File created:")
        print(f"  IDA Excel: {ida_excel.name}")
        print("="*70 + "\n")
        
        return ida_excel


# Standalone function for easy calling
def compile_ida1_bids(trading_date_str, upload_sql=False):
    """
    Compile IDA-1 adjustment bids
    
    Args:
        trading_date_str: Trading date in DD/MM/YYYY format
        upload_sql: Whether to upload to SQL (default: False)
    
    Returns:
        Path to IDA Excel file
    """
    compiler = IDA1BidCompiler()
    return compiler.run_ida1_compilation(trading_date_str, upload_sql=upload_sql)


if __name__ == "__main__":
    # Test
    files = compile_ida1_bids("23/01/2026")
    print(f"\nIDA-1 compilation successful!")
